#!/usr/bin/env python3
"""多线程喀什卫建购卡小工具 - PyQt5 版"""

import sys, os, time, random, csv, threading, re, base64
from io import BytesIO
import requests
from concurrent.futures import ThreadPoolExecutor, as_completed
from PyQt5.QtCore import Qt, QObject, pyqtSignal, QTimer
from PyQt5.QtGui import QFont, QColor, QBrush, QPalette, QIntValidator, QPixmap
from PyQt5.QtWidgets import (
    QApplication, QWidget, QVBoxLayout, QHBoxLayout, QDialog, QDialogButtonBox,
    QPushButton, QLabel, QProgressBar, QLineEdit,
    QTableWidget, QTableWidgetItem, QHeaderView,
    QAbstractItemView, QFileDialog, QMessageBox,
    QStatusBar, QFrame, QMenu, QAction,
)
from PyQt5.QtCore import Qt, QObject, pyqtSignal
from PyQt5.QtGui import QFont, QColor, QBrush, QPalette, QIntValidator

# ═══════════════════════════════════════════
#  超级鹰配置 — 修改成你的账号
# ═══════════════════════════════════════════
SUPER_EAGLE_USER = "alfaa3"
SUPER_EAGLE_PASS = "123456"
SUPER_EAGLE_SOFTID = "976381"
# ═══════════════════════════════════════════

# 参数名称 → ID 映射（全名匹配）
PARAM_MAP = {
    "临床、医技相关专项": "202610000035",
    "药学专项": "202610000036",
    "护理专项": "202610000034",
    "乡镇卫生院专项": "202610000037",
}

try:
    import openpyxl
except ImportError:
    openpyxl = None
try:
    import xlrd
except ImportError:
    xlrd = None

COLORS = {
    "成功": QColor("#1b5e20"),
    "失败": QColor("#b71c1c"),
    "已取消": QColor("#e65100"),
    "已导入": QColor("#0d47a1"),
}


class SignalBridge(QObject):
    progress = pyqtSignal(int)
    render = pyqtSignal()
    msg = pyqtSignal(str)
    done = pyqtSignal()
    error = pyqtSignal(str)
    enable = pyqtSignal(bool)
    show_qr = pyqtSignal(object, str, int, str)  # sess, out_trade_no, row, qr_url


class App(QWidget):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("多线程小工具")
        self.resize(940, 640)
        self.setMinimumSize(800, 520)

        self.rows = []
        self.has_data = False
        self.task_busy = False
        self.cancel = False

        self.sig = SignalBridge()
        self._build_ui()
        self.sig.progress.connect(self.progress.setValue)
        self.sig.render.connect(self._render)
        self.sig.msg.connect(self.lbl_prog.setText)
        self.sig.done.connect(self._task_end)
        self.sig.error.connect(lambda e: QMessageBox.critical(self, "错误", e))
        self.sig.enable.connect(self._update_btns)
        self.sig.show_qr.connect(self._show_qr_dialog)

    def _build_ui(self):
        self.setStyleSheet("""
            QWidget { font-family: "Helvetica Neue", Helvetica, Arial, sans-serif; }
        """)
        root = QVBoxLayout()
        root.setContentsMargins(0, 0, 0, 0)
        root.setSpacing(0)

        # ── 顶栏 ──
        top_frame = QFrame()
        top_frame.setStyleSheet("""
            QFrame { background: qlineargradient(x1:0, y1:0, x2:0, y2:1,
                stop:0 #f5f5f5, stop:1 #e8e8e8); border-bottom: 1px solid #ccc; }
        """)
        tl = QHBoxLayout(top_frame)
        tl.setContentsMargins(12, 8, 12, 8)
        tl.setSpacing(6)

        btn_s = "QPushButton{padding:7px 18px;border:none;border-radius:5px;font-size:13px;font-weight:bold;color:white;}"
        sep = lambda: (lambda s: (s.setFrameShape(QFrame.VLine), s.setStyleSheet("color:#ccc;"), tl.addWidget(s)))(QFrame())

        self.bi = QPushButton("📥 导入数据")
        self.bi.setStyleSheet(btn_s + "QPushButton{background:#4a86e8;}")
        self.bi.clicked.connect(self._import)
        tl.addWidget(self.bi)
        sep()

        self.ba = QPushButton("✅ 申请")
        self.ba.setStyleSheet(btn_s + "QPushButton{background:#fbbc04;}")
        self.ba.clicked.connect(self._apply)
        tl.addWidget(self.ba)

        self.bs = QPushButton("⏹ 停止")
        self.bs.setStyleSheet(btn_s + "QPushButton{background:#ea4335;}")
        self.bs.clicked.connect(self._stop)
        self.bs.setEnabled(False)
        tl.addWidget(self.bs)
        sep()

        self.bc = QPushButton("🗑 清空")
        self.bc.setStyleSheet(btn_s + "QPushButton{background:#888;}")
        self.bc.clicked.connect(self._clear)
        tl.addWidget(self.bc)

        self.bp = QPushButton("🔍 排查")
        self.bp.setStyleSheet(btn_s + "QPushButton{background:#9c27b0;}")
        self.bp.clicked.connect(self._filter_fail)
        self.bp.setEnabled(False)
        tl.addWidget(self.bp)

        self.be = QPushButton("📤 导出")
        self.be.setStyleSheet(btn_s + "QPushButton{background:#00bcd4;}")
        self.be.clicked.connect(self._export)
        tl.addWidget(self.be)
        tl.addStretch()

        tl.addWidget(QLabel("线程:", styleSheet="color:#555;font-size:13px;"))
        self.spin = QLineEdit("1")
        self.spin.setFixedWidth(50)
        self.spin.setMaxLength(2)
        self.spin.setValidator(QIntValidator(1, 64))
        self.spin.setStyleSheet("QLineEdit{padding:2px 6px;border:1px solid #bbb;border-radius:4px;background:white;font-size:13px;}")
        tl.addWidget(self.spin)
        root.addWidget(top_frame)

        # ── 表格 ──
        tf = QFrame()
        tfl = QVBoxLayout(tf)
        tfl.setContentsMargins(10, 6, 10, 4)

        self.table = QTableWidget()
        self.table.setColumnCount(7)
        self.table.setHorizontalHeaderLabels(["序号", "账号", "密码", "参数", "状态", "耗时", "备注"])
        self.table.setEditTriggers(QAbstractItemView.NoEditTriggers)
        self.table.setSelectionBehavior(QAbstractItemView.SelectRows)
        self.table.setSelectionMode(QAbstractItemView.SingleSelection)
        self.table.setAlternatingRowColors(True)
        self.table.verticalHeader().setVisible(False)
        self.table.setShowGrid(True)
        self.table.setGridStyle(Qt.SolidLine)
        self.table.horizontalHeader().setStretchLastSection(True)
        self.table.horizontalHeader().setSectionResizeMode(QHeaderView.Interactive)
        self.table.horizontalHeader().setFixedHeight(30)
        self.table.horizontalHeader().setStyleSheet("""
            QHeaderView::section {
                background: qlineargradient(x1:0,y1:0,x2:0,y2:1,stop:0#f0f0f0,stop:1#ddd);
                color:#333; font-weight:bold; font-size:12px; padding:4px 8px;
                border:none; border-right:1px solid #ccc; border-bottom:1px solid #ccc;
            }
        """)
        self.table.setColumnWidth(0, 45)
        self.table.setColumnWidth(1, 150)
        self.table.setColumnWidth(2, 120)
        self.table.setColumnWidth(3, 180)
        self.table.setColumnWidth(4, 80)
        self.table.setColumnWidth(5, 75)
        self.table.verticalHeader().setDefaultSectionSize(28)
        self.table.setStyleSheet("""
            QTableWidget { border:1px solid #ccc; border-radius:4px; background:white;
                alternate-background-color:#f5f7fa; gridline-color:#e0e0e0; }
            QTableWidget::item:selected { background:#cce5ff; color:black; }
        """)
        self.table.setContextMenuPolicy(Qt.CustomContextMenu)
        self.table.customContextMenuRequested.connect(self._table_menu)

        tfl.addWidget(self.table)
        root.addWidget(tf)

        # ── 进度 ──
        pf = QFrame()
        pl = QVBoxLayout(pf)
        pl.setContentsMargins(10, 0, 10, 2)
        self.progress = QProgressBar()
        self.progress.setTextVisible(False)
        self.progress.setFixedHeight(8)
        self.progress.setStyleSheet("""
            QProgressBar { border:1px solid #ddd; border-radius:4px; background:#eee; }
            QProgressBar::chunk {
                background: qlineargradient(x1:0,y1:0,x2:1,y2:0,stop:0#4a86e8,stop:1#34a853);
                border-radius:3px; }
        """)
        pl.addWidget(self.progress)
        self.lbl_prog = QLabel("就绪 — 点击「导入数据」选择文件")
        self.lbl_prog.setStyleSheet("color:#888; font-size:12px; padding:1px 0;")
        pl.addWidget(self.lbl_prog)
        root.addWidget(pf)

        # ── 状态栏 ──
        self.status = QStatusBar()
        self.status.setStyleSheet("QStatusBar{background:#f0f0f0;border-top:1px solid #d0d0d0;padding:2px 10px;font-size:12px;color:#666;}")
        self.status.showMessage("就绪 · 0 条数据")
        root.addWidget(self.status)
        self.setLayout(root)

        self._update_btns()

    def _update_btns(self):
        """按钮状态"""
        idle = not self.task_busy
        self.bi.setEnabled(not self.has_data and idle)
        self.ba.setEnabled(self.has_data and idle)
        self.bc.setEnabled(self.has_data and idle)
        self.bs.setEnabled(self.task_busy)
        # 排查: 有失败数据 + 空闲时启用
        has_fail = any(r["status"] == "失败" for r in self.rows) if self.rows else False
        self.bp.setEnabled(idle and has_fail)

    def _table_menu(self, pos):
        row = self.table.rowAt(pos.y())
        if row < 0:
            return
        self.table.selectRow(row)
        menu = QMenu(self)
        a1 = menu.addAction("🏙 申城")
        a2 = menu.addAction("💳 购卡")
        act = menu.exec_(self.table.viewport().mapToGlobal(pos))
        if act == a1:
            self._do_action(row, "申城")
        elif act == a2:
            self._do_action(row, "购卡")

    def _do_action(self, row, name):
        r = self.rows[row]
        if name == "购卡":
            threading.Thread(target=self._buy_card_for_row, args=(row,), daemon=True).start()
            return
        # 申城（简单模拟）
        r = self.rows[row]
        self.lbl_prog.setText(f"{name}: {r['account']}")
        time.sleep(0.3)
        r["status"] = "成功"
        r["cost"] = "0.3s"
        r["detail"] = f"{name}完成"
        self._render()

    # ═══════════════════ 购卡 ═══════════════════
    def _buy_card_for_row(self, row):
        """单行购卡流程"""
        r = self.rows[row]
        # 1. 检查备注列是否为"未购卡"
        if r["detail"] != "未购卡":
            self.sig.error.emit("此用户状态不是未购卡")
            return

        self.lbl_prog.setText(f"购卡: {r['account']}")
        account, password, param = r["account"], r["password"], r["param"]

        # 2. 登录
        sess, err = self._login_get_session(account, password)
        if err:
            r["status"] = "失败"
            r["detail"] = err
            self.sig.render.emit()
            return

        # 3. 获取课程ID
        param_id = PARAM_MAP.get(param, param)
        cw = sess.get(f"https://www.xjyxjyw.com/member/cw_info.do?id={param_id}", timeout=15)
        cids = re.findall(r'course_id=(\d+)', cw.text)
        if not cids:
            r["status"] = "失败"
            r["detail"] = "未找到课程"
            self.sig.render.emit()
            return
        course_id = cids[0]

        # 4. 访问购卡页 + 提交微信支付
        product_id = "2026049" if "乡镇卫生院专项" in param else "2026041"
        buy_url = f"https://www.xjyxjyw.com/member/myCard_buycard.do?product_id={product_id}&ids={course_id}&paycode=02&product_quantity=1"
        buy_resp = sess.get(buy_url, timeout=15)

        # 5. 解析表单字段
        form_fields = {}
        for name_field in ("WIDbody", "WIDshow_url", "out_trade_no", "WIDsubject", "WIDtotal_fee"):
            m = re.search(r'name=["\']' + name_field + r'["\'][^>]*value=["\']([^"\']+)["\']', buy_resp.text)
            if m:
                form_fields[name_field] = m.group(1)
            else:
                m = re.search(r'name=["\']' + name_field + r'["\'](?:[^>]*>)\s*([^<\s]+)', buy_resp.text)
                if m:
                    form_fields[name_field] = m.group(1)

        if "out_trade_no" not in form_fields:
            r["status"] = "失败"
            r["detail"] = "未找到订单号"
            self.sig.render.emit()
            return

        out_trade_no = form_fields["out_trade_no"]

        # 构建POST数据
        post_data = {
            "WIDbody": form_fields.get("WIDbody", ""),
            "WIDshow_url": form_fields.get("WIDshow_url", ""),
            "out_trade_no": out_trade_no,
            "WIDsubject": form_fields.get("WIDsubject", ""),
            "WIDtotal_fee": form_fields.get("WIDtotal_fee", ""),
        }
        # WIDsubject 和 WIDtotal_fee 各出现两次(名称/数量, 金额/方式)
        # 第二个 WIDsubject = 数量 1, 第二个 WIDtotal_fee = 微信
        post_data["WIDsubject"] = post_data["WIDsubject"] + "&WIDsubject=1"
        post_data["WIDtotal_fee"] = post_data["WIDtotal_fee"] + "&WIDtotal_fee=微信"

        # 6. 提交获取二维码页面
        wx_resp = sess.post("https://www.xjyxjyw.com/member/myCard_wxPay.do",
                            data=post_data, timeout=15,
                            headers={"Referer": buy_url, "Content-Type": "application/x-www-form-urlencoded"})

        # 7. 提取二维码 URL（JS 生成的微信支付链接）
        qr_url = ""
        # 从 JS 中提取 url 变量
        url_m = re.search(r'var\s+url\s*=\s*"([^"]+)"', wx_resp.text)
        if url_m:
            qr_url = url_m.group(1)
        else:
            url_m = re.search(r"var\s+url\s*=\s*'([^']+)'", wx_resp.text)
            if url_m:
                qr_url = url_m.group(1)

        # 8. 显示二维码弹窗（通过信号在主线程执行）
        self.sig.show_qr.emit(sess, out_trade_no, row, qr_url)

    def _login_get_session(self, account, password):
        """登录获取session，返回 (session, None) 或 (None, 错误信息)"""
        sess = requests.Session()
        sess.headers.update({"User-Agent": "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) "
                                           "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/148.0.0.0 Safari/537.36"})

        sess.get("https://www.xjyxjyw.com/mlogin.jsp", timeout=15)
        if self.cancel:
            return None, "已取消"

        ts = time.strftime("a%20b%20c%20d%20%H:%M:%S", time.localtime())
        cap = sess.get(f"https://www.xjyxjyw.com/image.jsp?date={ts}", timeout=15)
        if self.cancel:
            return None, "已取消"

        code = self._ocr(cap.content)
        if not code:
            return None, "验证码识别失败"
        if code.startswith("__ERR:"):
            return None, f"超级鹰:{code[6:]}"

        r = sess.post("https://www.xjyxjyw.com/member_login.do",
                      data={"member.loginname": account, "member.pwd": password, "ValidateCode": code},
                      timeout=15, allow_redirects=False)
        if self.cancel:
            return None, "已取消"

        if r.status_code not in (302, 303, 307):
            html = r.text
            if "验证码" in html and "错误" in html:
                return None, "验证码错误"
            elif "密码" in html and "错误" in html:
                return None, "密码错误"
            elif "没有您的信息" in html:
                return None, "无此账号"
            return None, "登录失败"

        sess.get("https://www.xjyxjyw.com/member_login.do", timeout=15)
        return sess, None

    def _show_qr_dialog(self, sess, out_trade_no, row, qr_url=""):
        """显示二维码弹窗 + 轮询支付状态（必须在主线程调用）"""
        dialog = QDialog(self)
        dialog.setWindowTitle("微信扫码支付")
        dialog.setFixedSize(320, 400)

        layout = QVBoxLayout(dialog)

        # 二维码图片
        qr_label = QLabel()
        qr_label.setAlignment(Qt.AlignCenter)
        qr_label.setMinimumSize(280, 280)
        if qr_url:
            # 用 qrcode 库生成二维码
            try:
                import qrcode
                qr_img = qrcode.make(qr_url)
                buf = BytesIO()
                qr_img.save(buf, format="PNG")
                pixmap = QPixmap()
                pixmap.loadFromData(buf.getvalue())
                qr_label.setPixmap(pixmap.scaled(280, 280, Qt.KeepAspectRatio, Qt.SmoothTransformation))
            except Exception:
                qr_label.setText(f"二维码生成失败\nURL: {qr_url[:50]}")
        else:
            qr_label.setText("未获取到支付链接")

        layout.addWidget(qr_label)

        info = QLabel(f"订单号: {out_trade_no}\n等待扫码支付...")
        info.setAlignment(Qt.AlignCenter)
        layout.addWidget(info)

        btn_close = QPushButton("关闭")
        btn_close.clicked.connect(dialog.reject)
        layout.addWidget(btn_close)

        # 轮询
        timer = QTimer()
        self._poll_count = 0

        def poll_state():
            self._poll_count += 1
            try:
                ts = int(time.time() * 1000)
                state_resp = sess.get(
                    f"https://www.xjyxjyw.com/member/myCard_state.do?out_trade_no={out_trade_no}&t={ts}",
                    timeout=10)
                text = state_resp.text.strip()
                print(f"[购卡] 第{self._poll_count}次轮询: {text}")
                if 'Satues:1' in text or '":1"' in text or "':1'" in text:
                    timer.stop()
                    r = self.rows[row]
                    r["status"] = "成功"
                    r["cost"] = f"{self._poll_count * 3}s"
                    r["detail"] = "已购卡"
                    self._render()
                    dialog.accept()
                    QMessageBox.information(self, "成功", "购卡成功！")
                    # 购卡成功后自动申请学分
                    threading.Thread(
                        target=self._apply_with_session,
                        args=(sess, row),
                        daemon=True
                    ).start()
            except Exception:
                pass
            if self._poll_count > 120:  # 6分钟超时
                timer.stop()
                dialog.reject()

        timer.timeout.connect(poll_state)
        timer.start(3000)

        dialog.exec_()
        if timer.isActive():
            timer.stop()

    def _apply_with_session(self, sess, row):
        """购卡成功后用已有session申请学分"""
        r = self.rows[row]
        self.lbl_prog.setText(f"申请学分: {r['account']}")
        param_id = PARAM_MAP.get(r["param"], r["param"])

        # 1. 获取课程列表
        cw = sess.get(f"https://www.xjyxjyw.com/member/cw_info.do?id={param_id}", timeout=15)
        course_ids = []
        for tr in re.findall(r'<tr[^>]*>(.*?)</tr>', cw.text, re.DOTALL):
            if "通过考试" in tr and "申请" in tr:
                m = re.search(r'course_id=(\d+)', tr)
                if m:
                    course_ids.append(m.group(1))

        if not course_ids:
            self.lbl_prog.setText("申请学分: 无可申请课程")
            return

        # 2. 逐个课程申请
        applied = 0
        for cid in course_ids:
            if self.cancel:
                return
            # 访问申请页获取卡信息
            ap = sess.get(f"https://www.xjyxjyw.com/member/apply_apply.do?course_id={cid}", timeout=15)
            for card_tr in re.findall(r'<tr[^>]*>(.*?)</tr>', ap.text, re.DOTALL):
                if "使用此卡申请" not in card_tr:
                    continue
                parts = re.findall(r'>([^<]+)<', card_tr)
                card_no = ""
                card_pwd = ""
                score = "0"
                for p in parts:
                    p = p.strip()
                    if p.isdigit() and len(p) >= 8 and not card_no:
                        card_no = p
                    elif p.isdigit() and len(p) >= 4 and card_no and not card_pwd:
                        card_pwd = p
                    else:
                        s = re.sub(r'[^\d.]', '', p)
                        if s:
                            try:
                                score = s
                            except ValueError:
                                pass
                if not card_no or not card_pwd:
                    continue
                try:
                    if float(re.sub(r'[^\d.]', '', score)) <= 0:
                        continue
                except ValueError:
                    continue

                # 提交申请
                apply_url = f"https://www.xjyxjyw.com/member/apply_applyCard.do?courseLog.cid={cid}&cardNo={card_no}&cardPasswd={card_pwd}"
                resp = sess.get(apply_url, timeout=15)
                if "申请成功" in resp.text:
                    applied += 1
                    break

        # 3. 更新状态
        r = self.rows[row]
        if applied > 0:
            r["status"] = "成功"
            r["detail"] = f"已购卡+申请{applied}门"
        else:
            r["detail"] = "已购卡(申请失败)"
        self._render()
        self.lbl_prog.setText(f"申请学分完成: {r['account']}")

    def _render(self):
        self.table.setRowCount(len(self.rows))
        for i, r in enumerate(self.rows):
            vals = [str(i+1), r["account"], r["password"], r["param"],
                    r["status"], r["cost"], r["detail"]]
            for j, v in enumerate(vals):
                item = QTableWidgetItem(v)
                item.setTextAlignment(Qt.AlignCenter if j in (0,4,5) else Qt.AlignLeft | Qt.AlignVCenter)
                if j == 4 and v in COLORS:
                    item.setForeground(QBrush(COLORS[v]))
                    f = item.font()
                    f.setBold(True)
                    item.setFont(f)
                self.table.setItem(i, j, item)
        self.status.showMessage(f"就绪 · {len(self.rows)} 条数据")

    # ═══════════════════ 导入 ═══════════════════
    def _import(self):
        path, _ = QFileDialog.getOpenFileName(self, "选择数据文件", "",
            "Excel/文本 (*.xlsx *.xls *.csv *.txt);;Excel (*.xlsx *.xls);;CSV (*.csv);;文本 (*.txt)")
        if not path:
            return
        threading.Thread(target=self._iw, args=(path,), daemon=True).start()

    def _iw(self, path):
        self.cancel = False
        self.sig.enable.emit(False)
        self.sig.msg.emit(f"导入中: {os.path.basename(path)}")
        ext = os.path.splitext(path)[1].lower()
        try:
            rec = self._px(path) if ext in (".xlsx", ".xls") else self._pt(path)
        except Exception as e:
            self.sig.error.emit(str(e))
            self.sig.enable.emit(True)
            return
        for i in range(len(rec)):
            time.sleep(0.005)
            self.sig.progress.emit(int((i+1)/len(rec)*100))
        self.rows = [dict(account=a, password=p, param=pm, status="已导入", cost="", detail="")
                      for a, p, pm in rec]
        self.has_data = True
        self.sig.render.emit()
        self.sig.progress.emit(100)
        self.sig.msg.emit(f"导入完成 · {len(rec)} 条")
        self.sig.enable.emit(True)

    def _px(self, path):
        ext = os.path.splitext(path)[1].lower()
        rec = []
        if ext == ".xlsx":
            if not openpyxl:
                raise Exception("需 openpyxl")
            wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
            ws = wb.active
            for row in ws.iter_rows(min_row=1, values_only=True):
                vs = [str(v).strip() if v is not None else "" for v in row]
                vs = [v for v in vs if v]
                if vs:
                    rec.append((vs[0], vs[1] if len(vs) > 1 else "", vs[2] if len(vs) > 2 else ""))
            wb.close()
        elif ext == ".xls":
            if not xlrd:
                raise Exception("需 xlrd")
            wb = xlrd.open_workbook(path)
            ws = wb.sheet_by_index(0)
            for r in range(ws.nrows):
                vs = []
                for c in range(ws.ncols):
                    v = ws.cell_value(r, c)
                    t = ws.cell_type(r, c)
                    if t == 0:
                        continue
                    if t == 2:
                        vs.append(str(int(v)) if v == int(v) else str(v))
                    else:
                        vs.append(str(v).strip())
                vs = [v.strip() for v in vs if v.strip()]
                if vs:
                    rec.append((vs[0], vs[1] if len(vs) > 1 else "", vs[2] if len(vs) > 2 else ""))
        return rec

    def _pt(self, path):
        with open(path, "r", encoding="utf-8", errors="replace") as f:
            raw = f.read()
        lines = raw.strip().splitlines()
        if not lines:
            return []
        try:
            rows = list(csv.reader(lines))
            if rows and len(rows[0]) >= 2:
                rec = []
                for r in rows:
                    if r and r[0].strip():
                        vs = [v.strip() for v in r]
                        rec.append((vs[0], vs[1] if len(vs) > 1 else "", vs[2] if len(vs) > 2 else ""))
                return rec
        except Exception:
            pass
        rec = []
        for line in lines:
            line = line.strip()
            if not line:
                continue
            for sep in ("\t", ",", "|", ";", "  "):
                if sep in line:
                    ps = [p.strip() for p in line.split(sep) if p.strip()]
                    rec.append((ps[0], ps[1] if len(ps) > 1 else "", ps[2] if len(ps) > 2 else ""))
                    break
            else:
                rec.append((line, "", ""))
        if not rec:
            rec = [(l.strip(), "", "") for l in lines if l.strip()]
        if len(rec) > 1 and any(k in rec[0][0].lower() for k in ("账号", "密码", "account", "序号")):
            rec = rec[1:]
        return rec

    # ═══════════════════ 申请 ═══════════════════
    def _apply(self):
        if not self.rows:
            QMessageBox.warning(self, "提示", "请先导入数据")
            return
        threading.Thread(target=self._apply_worker, daemon=True).start()

    def _apply_worker(self):
        """每个账号：登录 → 查课程 → 分析状态"""
        self.task_busy = True
        self.cancel = False
        self.progress.setValue(0)
        self.sig.enable.emit(False)
        n = int(self.spin.text() or 5)
        total = len(self.rows)
        done = [0]
        self.sig.msg.emit(f"申请中 · 线程: {n}")

        def work(idx, row):
            if self.cancel:
                return idx, "已取消", "", ""
            try:
                st, detail, cost = self._login_and_check(row["account"], row["password"], row["param"])
                return idx, st, cost, detail
            except Exception as e:
                return idx, "失败", "", str(e)[:80]

        with ThreadPoolExecutor(max_workers=n) as pool:
            fs = {pool.submit(work, i, r): i for i, r in enumerate(self.rows)}
            for f in as_completed(fs):
                if self.cancel:
                    for i, r in enumerate(self.rows):
                        if r["status"] in ("", "已导入"):
                            self.rows[i]["status"] = "已取消"
                    break
                idx, st, cost, detail = f.result()
                self.rows[idx]["status"] = st
                self.rows[idx]["cost"] = cost
                self.rows[idx]["detail"] = detail
                done[0] += 1
                self.sig.render.emit()
                if done[0] % 3 == 0 or done[0] == total:
                    self.sig.progress.emit(int(done[0] / total * 100))

        self.sig.render.emit()
        self.sig.progress.emit(100)
        ok = sum(1 for r in self.rows if r["status"] == "成功")
        fail = sum(1 for r in self.rows if r["status"] == "失败")
        can = sum(1 for r in self.rows if r["status"] == "已取消")
        pts = " / ".join(p for p in [f"{ok}成功" if ok else "", f"{fail}失败" if fail else "",
                                      f"{can}取消" if can else ""] if p)
        self.sig.msg.emit(f"申请完成 · {pts}")
        self.sig.done.emit()

    def _login_and_check(self, account, password, param):
        """登录 → 课程页 → 返回 (状态, 详情, 耗时)"""
        sess = requests.Session()
        sess.headers.update({"User-Agent": "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) "
                                           "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/148.0.0.0 Safari/537.36"})
        t0 = time.time()

        # 1. 获取 JSESSIONID
        sess.get("https://www.xjyxjyw.com/mlogin.jsp", timeout=15)
        if self.cancel:
            return "已取消", "", ""

        # 2. 验证码
        ts = time.strftime("a%20b%20c%20d%20%H:%M:%S", time.localtime())
        cap = sess.get(f"https://www.xjyxjyw.com/image.jsp?date={ts}", timeout=15)
        if self.cancel:
            return "已取消", "", ""

        # 3. 超级鹰识别
        code = self._ocr(cap.content)
        if not code:
            return "失败", "验证码识别失败", f"{time.time()-t0:.1f}s"
        if code.startswith("__ERR:"):
            return "失败", f"超级鹰:{code[6:]}", f"{time.time()-t0:.1f}s"
        if len(code) != 4:
            return "失败", f"验证码位数不对({code})", f"{time.time()-t0:.1f}s"

        # 4. 登录
        r = sess.post("https://www.xjyxjyw.com/member_login.do",
                      data={"member.loginname": account, "member.pwd": password, "ValidateCode": code},
                      timeout=15, allow_redirects=False)
        if self.cancel:
            return "已取消", "", ""

        elapsed = f"{time.time()-t0:.1f}s"

        # 5. 判断登录
        if r.status_code not in (302, 303, 307):
            html = r.text
            if "验证码" in html and "错误" in html:
                return "失败", "验证码错误", elapsed
            elif "密码" in html and "错误" in html:
                return "失败", "密码错误", elapsed
            elif "没有您的信息" in html:
                return "失败", "无此账号", elapsed
            return "失败", "登录失败", elapsed

        # 登录成功，跟随重定向
        sess.get("https://www.xjyxjyw.com/member_login.do", timeout=15)

        # 6. 检查IC卡
        acct_resp = sess.get("https://www.xjyxjyw.com/member/member_account.do", timeout=15)
        ic_match = re.search(r'name=["\']member\.icCard["\'][^>]*value=["\']([^"\']*)["\']', acct_resp.text)
        ic_card = ic_match.group(1).strip() if ic_match else ""
        print(f"[IC卡] {account} -> {ic_card}")
        if not ic_card:
            return "失败", "未绑定IC卡", f"{time.time()-t0:.1f}s"

        # 7. 课程页
        # 参数名全匹配 → ID
        param_id = PARAM_MAP.get(param, param)
        cw = sess.get(f"https://www.xjyxjyw.com/member/cw_info.do?id={param_id}", timeout=15)
        elapsed = f"{time.time()-t0:.1f}s"

        # 7. 解析课程 & 申请学分
        states = []
        course_ids = []  # 可申请的课程ID
        for tr in re.findall(r'<tr[^>]*>(.*?)</tr>', cw.text, re.DOTALL):
            if "已申请学分" in tr:
                states.append("done")
            elif "通过考试" in tr and "申请" in tr:
                states.append("can_apply")
                # 提取 course_id
                m = re.search(r'course_id=(\d+)', tr)
                if m:
                    course_ids.append(m.group(1))
            elif any(kw in tr for kw in ("未学习", "学习中", "待考试")):
                states.append("blocked")

        if not states:
            return "失败", "未找到课程数据", elapsed

        has_blocked = any(s == "blocked" for s in states)
        has_can_apply = any(s == "can_apply" for s in states)

        if has_blocked:
            return "失败", "学习未完成", elapsed

        if not has_can_apply:
            return "成功", "已全部完成", elapsed

        # 有可申请的课程 → 执行申请
        all_success = True
        has_no_card = False
        for cid in course_ids:
            # 访问申请页，获取卡信息
            apply_page = sess.get(
                f"https://www.xjyxjyw.com/member/apply_apply.do?course_id={cid}",
                timeout=15)
            if self.cancel:
                return "已取消", "", elapsed

            used_card = False
            for card_tr in re.findall(r'<tr[^>]*>(.*?)</tr>', apply_page.text, re.DOTALL):
                if "使用此卡申请" not in card_tr:
                    continue
                parts = re.findall(r'>([^<]+)<', card_tr)
                card_no = ""
                card_pwd = ""
                score = "0"
                for i, p in enumerate(parts):
                    p = p.strip()
                    if p.isdigit() and len(p) >= 8 and not card_no:
                        card_no = p
                    elif p.isdigit() and len(p) >= 4 and card_no and not card_pwd:
                        card_pwd = p
                    elif "".join(c for c in p if c.isdigit() or c == ".").replace(".", "").isdigit():
                        score = p
                if not card_no or not card_pwd:
                    continue
                try:
                    score_val = float(re.sub(r'[^\d.]', '', score))
                except ValueError:
                    score_val = 0
                if score_val <= 0:
                    continue

                apply_url = (
                    f"https://www.xjyxjyw.com/member/apply_applyCard.do"
                    f"?courseLog.cid={cid}&cardNo={card_no}&cardPasswd={card_pwd}"
                )
                apply_resp = sess.get(apply_url, timeout=15)
                if "申请成功" in apply_resp.text:
                    used_card = True
                    break

            if not used_card:
                has_no_card = True

        if has_no_card:
            return "失败", "未购卡", elapsed
        elif all_success:
            return "成功", "申请成功", elapsed
        else:
            return "成功", "申请成功", elapsed

    def _ocr(self, img_data):
        """超级鹰"""
        if SUPER_EAGLE_USER == "your_username":
            return ""
        # 保存图片用于调试
        debug_path = "/tmp/captcha_last.jpg"
        with open(debug_path, "wb") as f:
            f.write(img_data)
        r = requests.post("https://upload.chaojiying.net/Upload/Processing.php",
                          data={"user": SUPER_EAGLE_USER, "pass": SUPER_EAGLE_PASS,
                                "softid": SUPER_EAGLE_SOFTID, "codetype": "4004"},
                          files={"userfile": ("c.jpg", img_data, "image/jpeg")}, timeout=15)
        try:
            j = r.json()
            if j.get("err_no") == 0:
                return j.get("pic_str", "")
            else:
                # 返回错误码，让调用方显示具体原因
                return f"__ERR:{j.get('err_no')}:{j.get('err_str','')}"
        except Exception as e:
            return f"__ERR:PARSE:{e}"

    # ═══════════════════ 操作 ═══════════════════
    def _stop(self):
        self.cancel = True
        self.lbl_prog.setText("已停止")
        self._task_end()

    def _clear(self):
        self.rows.clear()
        self.has_data = False
        self.task_busy = False
        self.table.setRowCount(0)
        self.progress.setValue(0)
        self.lbl_prog.setText("就绪 — 点击「导入数据」选择文件")
        self.status.showMessage("就绪 · 0 条数据")
        self._update_btns()

    def _filter_fail(self):
        """排查：只保留失败的行"""
        self.rows = [r for r in self.rows if r["status"] == "失败"]
        self._render()
        self._update_btns()

    def _export(self):
        """导出表格数据到 Excel"""
        if not self.rows:
            QMessageBox.warning(self, "提示", "没有数据可导出")
            return
        path, _ = QFileDialog.getSaveFileName(
            self, "导出数据", "导出数据.xlsx",
            "Excel (*.xlsx);;CSV (*.csv)")
        if not path:
            return

        threading.Thread(target=self._export_worker, args=(path,), daemon=True).start()

    def _export_worker(self, path):
        try:
            ext = os.path.splitext(path)[1].lower()
            headers = ["序号", "账号", "密码", "参数", "状态", "耗时", "备注"]
            data = []
            for i, r in enumerate(self.rows, 1):
                data.append([str(i), r["account"], r["password"], r["param"],
                             r["status"], r["cost"], r["detail"]])

            if ext == ".csv":
                with open(path, "w", newline="", encoding="utf-8-sig") as f:
                    writer = csv.writer(f)
                    writer.writerow(headers)
                    writer.writerows(data)
            else:
                if openpyxl:
                    wb = openpyxl.Workbook()
                    ws = wb.active
                    ws.append(headers)
                    for row in data:
                        ws.append(row)
                    wb.save(path)
                else:
                    self.sig.error.emit("需要安装 openpyxl 才能导出 Excel")
                    return

            self.sig.msg.emit(f"导出完成: {os.path.basename(path)}")
        except Exception as e:
            self.sig.error.emit(f"导出失败: {str(e)}")

    def _task_end(self):
        self.task_busy = False
        self._update_btns()


if __name__ == "__main__":
    app = QApplication(sys.argv)
    app.setStyle("Fusion")
    pal = app.palette()
    for role, color in [(pal.Window, "#f0f0f0"), (pal.WindowText, "#1d1d1f"),
                        (pal.Base, "#ffffff"), (pal.AlternateBase, "#f5f7fa"),
                        (pal.Text, "#1d1d1f"), (pal.Button, "#e8e8e8"),
                        (pal.ButtonText, "#1d1d1f"), (pal.Highlight, "#4a86e8"),
                        (pal.HighlightedText, "#ffffff")]:
        pal.setColor(role, QColor(color))
    app.setPalette(pal)
    w = App()
    w.show()
    sys.exit(app.exec_())
